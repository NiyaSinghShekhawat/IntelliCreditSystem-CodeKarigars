# src/extractor.py
"""
Financial data extractor for IntelliCredit.

ARCHITECTURE:
  Primary  → Groq LLaMA 3.3 70B reads raw document text and returns JSON.
             This handles any column naming variation, multi-language labels,
             and non-standard layouts without brittle regex patterns.
  Fallback → Original regex/table logic runs if Groq fails or returns zeros,
             ensuring extraction always produces something even offline.
"""

from config import DEBUG_MODE, GROQ_API_KEY, GROQ_MODEL
from src.schemas import (
    ParsedDocument, GSTData, BankStatementData,
    ITRData, DocumentType
)
import re
import json
import sys
from pathlib import Path
from typing import Optional

ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))


# ─── GROQ EXTRACTION PROMPTS ─────────────────────────────────────────────────

_GST_PROMPT = """You are a financial data extraction engine for Indian GST documents.
Extract the following fields from the document text below.
Return ONLY a valid JSON object — no explanation, no markdown, no extra text.

Fields to extract:
- gstin: string (15-char GST identification number, e.g. "27AABCU9603R1ZX")
- company_name: string
- tax_period: string (e.g. "April 2023 - March 2024")
- turnover: number (Total Taxable Turnover from OUTWARD SUPPLIES section in INR — NOT invoice values from supplier rows, NOT ITC amounts. Look for "Total Taxable Turnover" or "Aggregate Turnover". If document is GSTR-2A with only supplier rows and no turnover row, use null.)
- igst: number (IGST on outward supplies in INR)
- cgst: number (CGST on outward supplies in INR)
- sgst: number (SGST on outward supplies in INR)
- total_tax: number (total GST paid in INR — sum of IGST+CGST+SGST on outward supplies)
- itc_claimed: number (TOTAL Input Tax Credit in INR. Priority: use "TOTAL ITC CLAIMED", "Net ITC Available", or "Grand Total ITC" row if present. If no grand total row exists but individual supplier ITC rows are present (GSTR-2A format), SUM all individual "ITC Available" values from supplier rows to get the total. Never use a single supplier row as the total.)
- filing_regular: boolean (true if filing appears regular/on-time)

Critical rules:
- All monetary values must be plain numbers (no commas, no currency symbols)
- If a value is not found, use null — do NOT guess or infer
- turnover: ONLY from "Total Taxable Turnover" or "Aggregate Turnover" row. GSTR-2A documents typically do NOT have a turnover field — return null if not present.
- itc_claimed: use GRAND TOTAL row only — never individual supplier rows
- Do not confuse invoice value with turnover — invoice values are per-supplier amounts

DOCUMENT TEXT:
{text}

JSON:"""

_BANK_PROMPT = """You are a financial data extraction engine for Indian bank statements.
Extract the following fields from the document text below.
Return ONLY a valid JSON object — no explanation, no markdown, no extra text.

Fields to extract:
- bank_name: string (name of the bank)
- account_number: string (account number, masked is fine)
- account_holder: string
- period_start: string (statement start date)
- period_end: string (statement end date)
- total_credits: number (sum of ALL credit/deposit transactions in INR)
- total_debits: number (sum of ALL debit/withdrawal transactions in INR)
- average_monthly_balance: number (average monthly balance in INR)
- closing_balance: number (final/closing balance in INR)
- emi_bounce_count: number (count of bounced EMIs, cheque returns, or NACH returns — 0 if none)
- monthly_credits: array of numbers (monthly credit totals if available, else empty array)
- monthly_debits: array of numbers (monthly debit totals if available, else empty array)
- monthly_balances: array of numbers (month-end balances if available, else empty array)

Rules:
- All monetary values must be plain numbers (no commas, no currency symbols)
- If a value is not found, use null
- total_credits = all money coming IN (salary credits, NEFT in, RTGS in, etc.)
- total_debits = all money going OUT (withdrawals, NEFT out, EMI, etc.)
- emi_bounce_count = count of "Return", "Bounce", "NACH Return", "ECS Return" entries

DOCUMENT TEXT:
{text}

JSON:"""

_ITR_PROMPT = """You are a financial data extraction engine for Indian ITR (Income Tax Return) and financial statements.
Extract the following fields from the document text below.
Return ONLY a valid JSON object — no explanation, no markdown, no extra text.

Fields to extract:
- pan: string (PAN number)
- assessment_year: string (e.g. "2023-24")
- gross_income: number (gross total income in INR)
- net_income: number (net profit after tax / PAT in INR)
- tax_paid: number (total tax paid in INR)
- tds: number (TDS amount in INR)
- net_worth: number (net worth = shareholders equity = paid-up capital + reserves in INR)
- long_term_debt: number (long-term borrowings / term loans in INR)
- short_term_debt: number (short-term borrowings / CC limits / working capital loans in INR)
- revenue: number (revenue from operations / net sales in INR)
- interest_expense: number (finance costs / interest on borrowings in INR)
- depreciation: number (depreciation and amortisation in INR)
- total_assets: number (total assets in INR)
- total_liabilities: number (total liabilities including equity in INR)

Rules:
- All monetary values must be plain numbers (no commas, no currency symbols)
- If a value is not found, use null
- net_worth = Shareholders Equity = Share Capital + Reserves and Surplus (NOT total equity + liabilities)
- long_term_debt = Term Loans + Debentures + Long-term Borrowings (anything > 1 year maturity)
- short_term_debt = Cash Credit + Working Capital Loan + Short-term Borrowings + Current portion of LTD
- revenue = Revenue from Operations (top line, before expenses)
- net_income = PAT (Profit After Tax) — can be negative if company is loss-making

DOCUMENT TEXT:
{text}

JSON:"""


class FinancialExtractor:
    """
    Extracts structured financial data from parsed documents.

    Uses Groq LLaMA as primary extractor (handles any column format),
    with regex/table scan as fallback (works offline).
    """

    def __init__(self):
        self._groq_client = None
        self._groq_available = False
        self._gemini_available = False
        self._gemini_model = None
        self._init_groq()
        self._init_gemini()

    def _init_groq(self):
        """Initialise Groq client for LLM extraction."""
        try:
            from groq import Groq
            from config import GROQ_API_KEYS
            if GROQ_API_KEYS:
                self._groq_keys = list(GROQ_API_KEYS)  # all available keys
                self._groq_client = Groq(api_key=self._groq_keys[0])
                self._groq_available = True
                print(f"Extractor: Groq LLM extraction enabled. "
                      f"{len(self._groq_keys)} key(s) available.")
            else:
                self._groq_keys = []
                print("Extractor: GROQ_API_KEY not set — using regex fallback.")
        except ImportError:
            self._groq_keys = []
            print("Extractor: groq package not installed — using regex fallback.")
        except Exception as e:
            self._groq_keys = []
            print(f"Extractor: Groq init failed ({e}) — using regex fallback.")

    def _init_gemini(self):
        """Initialise Gemini as secondary LLM backend (fallback when Groq exhausted).
        Supports both google-generativeai 0.8.x and google-genai 1.x SDKs."""
        try:
            from config import GEMINI_API_KEY
            if not GEMINI_API_KEY:
                print("Extractor: GEMINI_API_KEY not set — Gemini fallback disabled.")
                return

            self._gemini_is_new_sdk = False
            try:
                # Try old SDK first (google-generativeai 0.8.x)
                import google.genai as genai
                genai.configure(api_key=GEMINI_API_KEY)
                self._gemini_model = genai.GenerativeModel(
                    model_name="gemini-1.5-flash",
                    generation_config=genai.GenerationConfig(
                        temperature=0.0,
                        max_output_tokens=1024,
                    )
                )
            except (ImportError, AttributeError):
                # Fall back to new SDK (google-genai 1.x)
                import google.genai as genai
                self._gemini_model = genai.Client(api_key=GEMINI_API_KEY)
                self._gemini_is_new_sdk = True

            self._gemini_available = True
            print("Extractor: Gemini fallback enabled (gemini-1.5-flash).")
        except ImportError:
            print(
                "Extractor: google-generativeai not installed — Gemini fallback disabled.")
        except Exception as e:
            print(f"Extractor: Gemini init failed ({e}).")

    def _gemini_extract(self, prompt: str, doc_label: str) -> Optional[dict]:
        """
        Call Gemini 1.5 Flash as fallback when all Groq keys are exhausted.
        Uses identical prompts — output format is the same JSON structure.
        """
        if not self._gemini_available or not self._gemini_model:
            return None
        try:
            if getattr(self, "_gemini_is_new_sdk", False):
                response = self._gemini_model.models.generate_content(
                    model="gemini-1.5-flash", contents=prompt)
                raw = response.candidates[0].content.parts[0].text.strip()
            else:
                response = self._gemini_model.generate_content(prompt)
                raw = response.text.strip()

            # Strip markdown fences if present
            raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
            raw = re.sub(r'\s*```$', '', raw, flags=re.MULTILINE)
            raw = raw.strip()

            data = json.loads(raw)
            print(f"Gemini extracted {doc_label} successfully.")
            return data

        except json.JSONDecodeError as e:
            print(f"Gemini JSON parse error for {doc_label}: {e}")
            return None
        except Exception as e:
            print(f"Gemini extraction failed for {doc_label}: {e}")
            return None

    def _llm_extract(self, prompt: str, doc_label: str) -> Optional[dict]:
        """
        Unified LLM extraction entry point.
        Priority: Groq (with key rotation) → Gemini → None (regex/openpyxl fallback)
        """
        # Try Groq first
        result = self._groq_extract(prompt, doc_label)
        if result is not None:
            return result

        # Groq exhausted or unavailable — try Gemini
        if self._gemini_available:
            print(f"Groq unavailable for {doc_label} — trying Gemini...")
            result = self._gemini_extract(prompt, doc_label)
            if result is not None:
                return result

        return None

    def _groq_extract(self, prompt: str, doc_label: str) -> Optional[dict]:
        """
        Call Groq LLM with an extraction prompt.
        Automatically rotates to the next API key on rate limit (429) errors.
        Returns parsed JSON dict or None on all failures.
        """
        if not self._groq_available:
            return None

        from groq import Groq

        keys_to_try = list(getattr(self, "_groq_keys", []))
        if not keys_to_try:
            return None

        for i, key in enumerate(keys_to_try):
            try:
                client = Groq(api_key=key)
                response = client.chat.completions.create(
                    model=GROQ_MODEL,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.0,
                    max_tokens=1024,
                )
                raw = response.choices[0].message.content.strip()

                # Strip markdown fences if present
                raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
                raw = re.sub(r'\s*```$', '', raw, flags=re.MULTILINE)
                raw = raw.strip()

                data = json.loads(raw)
                if DEBUG_MODE:
                    print(f"Groq extracted {doc_label}: {data}")
                return data

            except json.JSONDecodeError as e:
                print(f"Groq JSON parse error for {doc_label}: {e}")
                return None  # JSON error won't be fixed by rotating keys

            except Exception as e:
                err_str = str(e)
                if "429" in err_str or "rate_limit" in err_str.lower():
                    if i < len(keys_to_try) - 1:
                        print(
                            f"Groq key {i+1} rate limited — rotating to key {i+2}...")
                        continue
                    else:
                        print(
                            f"Groq: all {len(keys_to_try)} key(s) exhausted — using fallback.")
                        return None
                print(f"Groq extraction failed for {doc_label}: {e}")
                return None

        return None

    def _g(self, data: dict, key: str, default=0.0):
        """Safe getter — returns default if key missing or null."""
        val = data.get(key)
        if val is None:
            return default
        try:
            if isinstance(default, bool):
                return bool(val)
            if isinstance(default, str):
                return str(val) if val else default
            return float(val)
        except (ValueError, TypeError):
            return default

    def _truncate_text(self, text: str, max_chars: int = 6000) -> str:
        """
        Truncate to max_chars for Groq prompt.
        Takes first 4000 + last 2000 chars to capture headers and summary totals.
        """
        if len(text) <= max_chars:
            return text
        return text[:4000] + "\n...[middle truncated]...\n" + text[-2000:]

    # ─── GST EXTRACTION ──────────────────────────────────────────────────────

    def extract_gst(self, parsed: ParsedDocument) -> GSTData:
        """Extract GST return data. Groq primary, xlsx fallback always runs for xlsx files."""
        gst = GSTData()

        if self._groq_available or self._gemini_available:
            prompt = _GST_PROMPT.format(
                text=self._truncate_text(parsed.raw_text))
            data = self._llm_extract(prompt, "GST")
            if data:
                gst.gstin = self._g(data, "gstin", "")
                gst.company_name = self._g(data, "company_name", "")
                gst.tax_period = self._g(data, "tax_period", "")
                gst.turnover = self._g(data, "turnover", 0.0)
                gst.igst = self._g(data, "igst", 0.0)
                gst.cgst = self._g(data, "cgst", 0.0)
                gst.sgst = self._g(data, "sgst", 0.0)
                gst.itc_claimed = self._g(data, "itc_claimed", 0.0)
                gst.filing_regular = bool(data.get("filing_regular", True))
                gst.total_tax = (self._g(data, "total_tax", 0.0)
                                 or gst.igst + gst.cgst + gst.sgst)
                if DEBUG_MODE:
                    print(
                        f"GST via Groq: turnover={gst.turnover}, itc={gst.itc_claimed}")

        # Always run xlsx fallback for xlsx files — Groq often hallucinated the
        # wrong ITC (e.g. returns 3B value for 2A file). Xlsx fallback reads
        # the actual column values directly so it's always more reliable.
        if parsed.source_file and parsed.source_file.lower().endswith(('.xlsx', '.xls')):
            gst = self._extract_gst_itc_from_xlsx(parsed.source_file, gst)
        elif gst.turnover == 0 and gst.itc_claimed == 0:
            print("GST: Groq returned zeros — running regex fallback.")
            return self._extract_gst_regex(parsed)

        print(
            f"GST final: turnover={gst.turnover}, itc={gst.itc_claimed}, gstin={gst.gstin}")
        return gst

    def _extract_gst_regex(self, parsed: ParsedDocument) -> GSTData:
        text = parsed.raw_text
        gst = GSTData()

        m = re.search(
            r'\b([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1})\b', text)
        if m:
            gst.gstin = m.group(1)

        m = re.search(
            r'(?:Legal Name|Trade Name|Business Name)[:\s]+([A-Z][A-Za-z\s&.,-]{3,60})', text, re.IGNORECASE)
        if m:
            gst.company_name = m.group(1).strip()

        m = re.search(
            r'(?:Total Turnover|Aggregate Turnover|Total Value|Turnover)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
        if m:
            gst.turnover = self._parse_amount(m.group(1))

        for pattern, attr in [
            (r'IGST[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)', 'igst'),
            (r'CGST[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)', 'cgst'),
            (r'SGST[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)', 'sgst'),
        ]:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                setattr(gst, attr, self._parse_amount(m.group(1)))

        gst.total_tax = gst.igst + gst.cgst + gst.sgst

        m = re.search(
            r'(?:TOTAL\s+ITC\s+CLAIMED|TOTAL\s+ITC|Grand\s+Total\s+ITC|Net\s+ITC\s+Available)'
            r'[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
        if m:
            gst.itc_claimed = self._parse_amount(m.group(1))
        else:
            m = re.search(
                r'(?:ITC|Input Tax Credit)\s+(?:Claimed|Available|Credit)'
                r'[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
            if m:
                gst.itc_claimed = self._parse_amount(m.group(1))

        gst = self._extract_gst_from_tables(gst, parsed.tables)
        return gst

    def _extract_gst_from_tables(self, gst: GSTData, tables) -> GSTData:
        for table in tables:
            for row in table.rows:
                row_str = " ".join(str(c) for c in row).lower()

                if ("turnover" in row_str or "total value" in row_str) and gst.turnover == 0:
                    for cell in row:
                        a = self._parse_amount(str(cell))
                        if a > 0:
                            gst.turnover = a
                            break

                for tax_attr in ['igst', 'cgst', 'sgst']:
                    if tax_attr in row_str and getattr(gst, tax_attr) == 0:
                        for cell in row:
                            a = self._parse_amount(str(cell))
                            if a > 0:
                                setattr(gst, tax_attr, a)
                                break

                is_itc_total = (
                    ("itc" in row_str or "input tax credit" in row_str) and
                    any(kw in row_str for kw in [
                        "total", "grand", "net", "available", "claimed"])
                )
                if is_itc_total and gst.itc_claimed == 0:
                    amounts = [self._parse_amount(str(c)) for c in row]
                    valid = [a for a in amounts if a > 0]
                    if valid:
                        gst.itc_claimed = max(valid)

        if gst.itc_claimed == 0:
            for table in tables:
                itc_col_idx = None
                header_row = None
                for i, row in enumerate(table.rows):
                    row_str = " ".join(str(c) for c in row).lower()
                    if "itc" in row_str or "input tax" in row_str:
                        header_row = i
                        for j, cell in enumerate(row):
                            if "itc" in str(cell).lower():
                                itc_col_idx = j
                                break
                        break
                if header_row is not None and itc_col_idx is not None:
                    total = sum(
                        self._parse_amount(str(row[itc_col_idx]))
                        for row in table.rows[header_row + 1:]
                        if itc_col_idx < len(row)
                    )
                    if total > 0:
                        gst.itc_claimed = total

        return gst

    def _extract_gst_itc_from_xlsx(self, file_path: str, gst: "GSTData") -> "GSTData":
        """Direct openpyxl scan: sum ITC column from GSTR-2A supplier rows,
        or read TOTAL ITC CLAIMED row from GSTR-3B."""
        try:
            import openpyxl
            from pathlib import Path as _Path
            import tempfile as _tempfile

            if not file_path:
                return gst
            p = _Path(file_path)
            if not p.is_absolute():
                p = _Path(_tempfile.gettempdir()) / p.name
            if not p.exists() or not str(p).lower().endswith(('.xlsx', '.xls')):
                return gst

            wb = openpyxl.load_workbook(str(p), data_only=True)

            def to_float(cell):
                if cell is None:
                    return None
                if isinstance(cell, (int, float)):
                    return float(cell) if float(cell) > 0 else None
                if isinstance(cell, str):
                    cleaned = cell.replace(',', '').replace(
                        '₹', '').replace('Rs.', '').strip()
                    try:
                        v = float(cleaned)
                        return v if v > 0 else None
                    except ValueError:
                        return None
                return None

            # Keywords — 3B grand total row (only present in 3B files)
            total_itc_3b_kw = ['total itc claimed',
                               'net itc available', 'grand total itc']
            # ITC column header keywords (present in 2A supplier table)
            itc_col_kw = ['itc available', 'itc (₹)', 'input tax credit']
            # Rows to SKIP — reconciliation notes embedded in 2A that reference 3B
            skip_row_kw = ['itc available as per gstr-2a',
                           'difference / variance', 'variance (a - b)',
                           'reconciliation note', 'timing difference']
            turnover_kw = ['total taxable turnover', 'aggregate turnover',
                           'taxable turnover']

            for sh in wb.sheetnames:
                ws = wb[sh]
                rows = list(ws.iter_rows(values_only=True))

                # Step 1: Try 3B-style grand total row
                # Skip any row that looks like a reconciliation note referencing 3B
                for row in rows:
                    row_text = ' '.join(str(c).lower()
                                        for c in row if c is not None)
                    if any(skip in row_text for skip in skip_row_kw):
                        continue
                    if any(kw in row_text for kw in total_itc_3b_kw):
                        for cell in row:
                            v = to_float(cell)
                            if v and v > 10000:
                                gst.itc_claimed = v
                                print(
                                    f"GST xlsx: grand_total row → itc={v:,.0f}")
                                break
                    if gst.itc_claimed > 0:
                        break

                # Step 2: Find ITC column header and sum supplier rows (2A style)
                # Always do this for 2A files — overrides step 1 if sum differs
                header_row_idx = None
                itc_col_idx = None
                for i, row in enumerate(rows):
                    row_text = ' '.join(str(c).lower()
                                        for c in row if c is not None)
                    if any(kw in row_text for kw in itc_col_kw):
                        header_row_idx = i
                        for j, cell in enumerate(row):
                            if cell and any(kw in str(cell).lower() for kw in itc_col_kw):
                                itc_col_idx = j
                                break
                        break

                if header_row_idx is not None and itc_col_idx is not None:
                    itc_sum = 0.0
                    for row in rows[header_row_idx + 1:]:
                        row_text = ' '.join(str(c).lower()
                                            for c in row if c is not None)
                        # Skip summary/reconciliation rows — only count supplier data rows
                        if any(skip in row_text for skip in skip_row_kw):
                            continue
                        if any(kw in row_text for kw in ['total', 'grand total', 'net itc']):
                            continue
                        if itc_col_idx < len(row):
                            v = to_float(row[itc_col_idx])
                            if v and v > 1000:
                                itc_sum += v
                    if itc_sum > 0:
                        gst.itc_claimed = itc_sum
                        print(f"GST xlsx: supplier column sum → itc={itc_sum:,.0f} "
                              f"(header_row={header_row_idx}, col={itc_col_idx})")

                # Step 3: Extract turnover if missing
                if gst.turnover == 0:
                    for row in rows:
                        row_text = ' '.join(str(c).lower()
                                            for c in row if c is not None)
                        if any(kw in row_text for kw in turnover_kw):
                            for cell in row:
                                v = to_float(cell)
                                if v and v > 100000:
                                    gst.turnover = v
                                    break
                        if gst.turnover > 0:
                            break

            print(
                f"GST xlsx fallback: itc={gst.itc_claimed:,.0f}, turnover={gst.turnover:,.0f}")

        except Exception as e:
            print(f"GST xlsx fallback error: {e}")
        return gst

    # ─── BANK STATEMENT EXTRACTION ───────────────────────────────────────────

    def extract_bank(self, parsed: ParsedDocument) -> BankStatementData:
        """Extract bank statement data. Groq primary, regex fallback."""
        if self._groq_available or self._gemini_available:
            prompt = _BANK_PROMPT.format(
                text=self._truncate_text(parsed.raw_text))
            data = self._llm_extract(prompt, "Bank")
            if data:
                mc = data.get("monthly_credits")
                md = data.get("monthly_debits")
                mb = data.get("monthly_balances")
                monthly_credits = [float(x) for x in mc if x is not None] if isinstance(
                    mc, list) else []
                monthly_debits = [float(x) for x in md if x is not None] if isinstance(
                    md, list) else []
                monthly_balances = [float(x) for x in mb if x is not None] if isinstance(
                    mb, list) else []

                avg_bal = self._g(data, "average_monthly_balance", 0.0)
                if avg_bal == 0 and monthly_balances:
                    avg_bal = sum(monthly_balances) / len(monthly_balances)

                total_credits = self._g(data, "total_credits", 0.0)
                total_debits = self._g(data, "total_debits", 0.0)

                if total_credits > 0 or total_debits > 0:
                    # Core fields supported by all schema versions
                    kwargs = dict(
                        bank_name=self._g(data, "bank_name", ""),
                        account_number=self._g(data, "account_number", ""),
                        total_credits=total_credits,
                        total_debits=total_debits,
                        average_monthly_balance=avg_bal,
                        closing_balance=self._g(data, "closing_balance", 0.0),
                        emi_bounce_count=int(
                            self._g(data, "emi_bounce_count", 0)),
                        monthly_credits=monthly_credits,
                        monthly_debits=monthly_debits,
                        monthly_balances=monthly_balances,
                    )
                    # New fields — silently skip if schema is old/cached
                    try:
                        bank = BankStatementData(
                            **kwargs,
                            account_holder=self._g(data, "account_holder", ""),
                            period_start=self._g(data, "period_start", ""),
                            period_end=self._g(data, "period_end", ""),
                        )
                    except Exception:
                        bank = BankStatementData(**kwargs)

                    # Always override avg_bal with xlsx summary row — Groq averages
                    # all running balance column values which gives a wrong inflated figure.
                    # The xlsx fallback reads the actual "Average Daily Balance" summary row.
                    if parsed.source_file and parsed.source_file.lower().endswith(('.xlsx', '.xls')):
                        bank = self._extract_bank_balance_from_xlsx(
                            parsed.source_file, bank)

                    if DEBUG_MODE:
                        print(f"Bank via Groq: credits={bank.total_credits}, "
                              f"debits={bank.total_debits}, "
                              f"avg_bal={bank.average_monthly_balance}")
                    return bank

        print("Bank: Groq returned zeros — running regex fallback.")
        return self._extract_bank_regex(parsed)

    def _extract_bank_regex(self, parsed: ParsedDocument) -> BankStatementData:
        text = parsed.raw_text
        bank = BankStatementData()

        SEP = r'(?:[^0-9\n]*?\s{2,}|\s*[:(]\s*)[₹(]*'
        for pattern, attr in [
            (r'(?:Bank Name|Bank)\s*[:\s]\s*([A-Za-z\s]+(?:Bank|Financial))', 'bank_name'),
            (r'A/C\s*No[:\s.]*\s*([0-9X*]{6,20})|Account\s*(?:No|Number|#)[:\s]+([0-9X*]{6,20})', 'account_number'),
            # XLSX summary layout: "Average Monthly Credits (₹)  1,03,58,333"
            (r'Average Monthly Credits' + SEP +
             r'([\d,]+\.?\d*)', 'total_credits'),
            (r'Average Monthly Debits' + SEP +
             r'([\d,]+\.?\d*)', 'total_debits'),
            # Also handle "Total Credits / Total Deposits" labels
            (r'Total\s+(?:Credits?|Deposits?)' +
             SEP + r'([\d,]+\.?\d*)', 'total_credits'),
            (r'Total\s+(?:Debits?|Withdrawals?)' +
             SEP + r'([\d,]+\.?\d*)', 'total_debits'),
            (r'Average\s+(?:Daily|Monthly)\s+Balance' + SEP +
             r'([\d,]+\.?\d*)', 'average_monthly_balance'),
            (r'Closing\s+Balance' + SEP +
             r'([\d,]+\.?\d*)', 'closing_balance'),
        ]:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                # Use first non-None group (handles patterns with 2 capture groups)
                val = next((g for g in m.groups() if g is not None), None)
                if not val:
                    continue
                if attr in ('bank_name', 'account_number'):
                    if not getattr(bank, attr, ''):
                        setattr(bank, attr, val.strip())
                else:
                    if getattr(bank, attr, 0) == 0:
                        setattr(bank, attr, self._parse_amount(val))

        # Bounce count: "EMI/Loan Bounces (6 months)  0" → capture 0, not 6
        bounce_m = re.search(
            r'EMI.{0,20}Bounces\s*\([^)]*\)\s*(\d+)', text, re.IGNORECASE)
        if bounce_m:
            bank.emi_bounce_count = int(bounce_m.group(1))
        else:
            bank.emi_bounce_count = len(re.findall(
                r'(?:EMI\s+Bounce|Cheque\s+Return|NACH\s+Return|ECS\s+Return)',
                text, re.IGNORECASE))

        bank = self._extract_bank_from_tables(bank, parsed.tables)

        # Always run xlsx fallback for xlsx files — reads the correct summary
        # row ("Average Daily Balance") instead of averaging all running balances
        print(f"Bank regex result: avg_bal={bank.average_monthly_balance}, "
              f"credits={bank.total_credits}, debits={bank.total_debits}")
        if parsed.source_file and parsed.source_file.lower().endswith(('.xlsx', '.xls')):
            bank = self._extract_bank_balance_from_xlsx(
                parsed.source_file, bank)

        return bank

    def _detect_bank_header(self, header_row: list) -> dict:
        """Map column indices to semantic names by fuzzy header matching."""
        mapping = {}
        for i, cell in enumerate(header_row):
            cell_l = str(cell).lower().strip()
            if any(kw in cell_l for kw in ['credit', 'deposit', 'receipts']) and 'credit' not in mapping:
                mapping['credit'] = i
            if any(kw in cell_l for kw in ['debit', 'withdrawal', 'payments']) and 'debit' not in mapping:
                mapping['debit'] = i
            if any(kw in cell_l for kw in ['balance', 'bal']) and 'balance' not in mapping:
                mapping['balance'] = i
        return mapping

    def _extract_bank_from_tables(self, bank: BankStatementData, tables) -> BankStatementData:
        all_credits, all_debits, all_balances = [], [], []

        for table in tables:
            if not table.rows:
                continue

            header_map = {}
            header_row_idx = 0
            for i, row in enumerate(table.rows[:5]):
                hm = self._detect_bank_header(row)
                if len(hm) >= 2:
                    header_map = hm
                    header_row_idx = i
                    break

            if not header_map:
                for row in table.rows:
                    row_str = " ".join(str(c) for c in row).lower()
                    if "total" in row_str:
                        amounts = [self._parse_amount(
                            str(c)) for c in row if self._parse_amount(str(c)) > 1000]
                        if len(amounts) >= 2:
                            if bank.total_credits == 0:
                                bank.total_credits = amounts[0]
                            if bank.total_debits == 0:
                                bank.total_debits = amounts[1]
                continue

            for row in table.rows[header_row_idx + 1:]:
                max_idx = max(header_map.values(), default=0)
                if len(row) <= max_idx:
                    continue
                if 'credit' in header_map:
                    v = self._parse_amount(str(row[header_map['credit']]))
                    if v > 0:
                        all_credits.append(v)
                if 'debit' in header_map:
                    v = self._parse_amount(str(row[header_map['debit']]))
                    if v > 0:
                        all_debits.append(v)
                if 'balance' in header_map:
                    v = self._parse_amount(str(row[header_map['balance']]))
                    if v > 0:
                        all_balances.append(v)

        if all_credits and bank.total_credits == 0:
            bank.total_credits = sum(all_credits)
        if all_debits and bank.total_debits == 0:
            bank.total_debits = sum(all_debits)
        if all_balances:
            bank.monthly_balances = all_balances
            if bank.average_monthly_balance == 0:
                bank.average_monthly_balance = sum(
                    all_balances) / len(all_balances)

        return bank

    def _extract_bank_balance_from_xlsx(self, file_path: str, bank: "BankStatementData") -> "BankStatementData":
        """Direct openpyxl scan for average monthly balance when regex/Groq miss it."""
        try:
            import openpyxl
            from pathlib import Path as _Path
            import tempfile as _tempfile

            if not file_path:
                return bank
            p = _Path(file_path)
            if not p.is_absolute():
                p = _Path(_tempfile.gettempdir()) / p.name
            if not p.exists() or not str(p).lower().endswith(('.xlsx', '.xls')):
                return bank

            wb = openpyxl.load_workbook(str(p), data_only=True)

            def to_float(cell):
                """Parse numeric cell — handles both real numbers and
                Indian comma-formatted strings like '1,18,50,000'."""
                if cell is None:
                    return None
                if isinstance(cell, (int, float)):
                    return float(cell)
                if isinstance(cell, str):
                    cleaned = cell.replace(',', '').replace(
                        '₹', '').replace('Rs.', '').strip()
                    try:
                        return float(cleaned)
                    except ValueError:
                        return None
                return None

            # Keywords for summary rows
            avg_daily_kw = ['average daily balance', 'avg daily balance']
            avg_monthly_kw = ['average monthly balance', 'avg monthly balance',
                              'average monthly credits', 'avg monthly credits']
            avg_credit_kw = ['average monthly credits', 'avg monthly credits']
            avg_debit_kw = ['average monthly debits', 'avg monthly debits']
            total_credit_kw = ['total credits', 'total deposits']
            total_debit_kw = ['total debits', 'total withdrawals']
            month_end_kw = ['month-end balance',
                            'closing balance', 'opening balance']

            month_end_balances = []
            all_credits = []
            all_debits = []

            for sh in wb.sheetnames:
                ws = wb[sh]
                rows = list(ws.iter_rows(values_only=True))

                for i, row in enumerate(rows):
                    row_text = ' '.join(str(c).lower()
                                        for c in row if c is not None)
                    if not row_text.strip():
                        continue

                    # ── Summary row: Average Daily Balance ───────────────────
                    # Always override — summary row is more accurate than
                    # averaging running balances from transaction ledger
                    if any(kw in row_text for kw in avg_daily_kw):
                        for cell in row:
                            v = to_float(cell)
                            if v and v > 10000:
                                bank.average_monthly_balance = v
                                break

                    # ── Summary row: Average Monthly Credits ─────────────────
                    if bank.total_credits == 0:
                        if any(kw in row_text for kw in avg_credit_kw):
                            for cell in row:
                                v = to_float(cell)
                                if v and v > 10000:
                                    # avg monthly * 6 months = total
                                    bank.total_credits = v * 6
                                    break

                    # ── Summary row: Average Monthly Debits ──────────────────
                    if bank.total_debits == 0:
                        if any(kw in row_text for kw in avg_debit_kw):
                            for cell in row:
                                v = to_float(cell)
                                if v and v > 10000:
                                    bank.total_debits = v * 6
                                    break

                    # ── Transaction ledger: collect month-end balances ────────
                    if any(kw in row_text for kw in month_end_kw):
                        for cell in row:
                            v = to_float(cell)
                            if v and v > 100000:
                                month_end_balances.append(v)
                                break

                    # ── Transaction ledger: sum credits and debits ────────────
                    # Row format: Date | Desc | Ref | Debit | Credit | Balance
                    if len(row) >= 6:
                        debit_cell = row[3]
                        credit_cell = row[4]
                        debit_v = to_float(debit_cell)
                        credit_v = to_float(credit_cell)
                        if credit_v and credit_v > 100000:
                            all_credits.append(credit_v)
                        if debit_v and debit_v > 100000:
                            all_debits.append(debit_v)

            # Derive average monthly balance from month-end balances if not set
            if bank.average_monthly_balance == 0 and month_end_balances:
                bank.average_monthly_balance = round(
                    sum(month_end_balances) / len(month_end_balances), 2
                )

            # Derive total credits/debits from ledger if summary rows weren't found
            if bank.total_credits == 0 and all_credits:
                bank.total_credits = round(sum(all_credits), 2)
            if bank.total_debits == 0 and all_debits:
                bank.total_debits = round(sum(all_debits), 2)

            # EMI bounce count — always re-scan xlsx for accuracy.
            # Groq tends to undercount; xlsx ledger has the full picture.
            # Strategy 1: read the summary row "EMI/Loan Bounces (6 months)  N"
            # Strategy 2: count ledger rows that contain bounce/return keywords
            # Take the maximum of Groq value and xlsx-derived value.
            bounce_summary_kw = ['emi', 'loan bounce', 'bounces']
            xlsx_bounce_from_summary = 0
            xlsx_bounce_from_ledger = 0
            bounce_ledger_kw = ['bounce', 'nach return', 'ecs return', 'dishonour',
                                'cheque return', 'nach bounce']
            for sh in wb.sheetnames:
                ws = wb[sh]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' '.join(str(c).lower()
                                        for c in row if c is not None)
                    if not row_text.strip():
                        continue
                    # Summary row: "EMI/Loan Bounces (6 months)   9"
                    if 'bounce' in row_text and ('month' in row_text or 'emi' in row_text):
                        for cell in row:
                            try:
                                v = int(float(cell))
                                if 0 < v < 100:   # sanity: bounce counts 1-99
                                    xlsx_bounce_from_summary = max(
                                        xlsx_bounce_from_summary, v)
                            except (TypeError, ValueError):
                                continue
                    # Ledger rows with bounce/return in remarks
                    elif any(kw in row_text for kw in bounce_ledger_kw):
                        # Skip the summary/header rows
                        if 'month' not in row_text and 'summary' not in row_text:
                            xlsx_bounce_from_ledger += 1
            xlsx_bounce = max(xlsx_bounce_from_summary,
                              xlsx_bounce_from_ledger)
            if xlsx_bounce > bank.emi_bounce_count:
                print(f"Bank xlsx: bounce override {bank.emi_bounce_count} → {xlsx_bounce} "
                      f"(summary={xlsx_bounce_from_summary}, ledger={xlsx_bounce_from_ledger})")
                bank.emi_bounce_count = xlsx_bounce

            print(f"Bank xlsx fallback: avg_bal={bank.average_monthly_balance:,.0f}, "
                  f"credits={bank.total_credits:,.0f}, debits={bank.total_debits:,.0f}, "
                  f"bounces={bank.emi_bounce_count}")

        except Exception as e:
            print(f"Bank xlsx fallback error: {e}")
        return bank

    # ─── ITR EXTRACTION ──────────────────────────────────────────────────────

    def extract_itr(self, parsed: ParsedDocument) -> ITRData:
        """Extract ITR / financial statement data. Groq primary, regex fallback."""
        itr = ITRData()

        if self._groq_available or self._gemini_available:
            prompt = _ITR_PROMPT.format(
                text=self._truncate_text(parsed.raw_text))
            data = self._llm_extract(prompt, "ITR")
            if data:
                itr.pan = self._g(data, "pan", "")
                itr.assessment_year = self._g(data, "assessment_year", "")
                itr.gross_income = self._g(data, "gross_income", 0.0)
                itr.net_income = self._g(data, "net_income", 0.0)
                itr.tax_paid = self._g(data, "tax_paid", 0.0)
                itr.tds = self._g(data, "tds", 0.0)
                itr.net_worth = self._g(data, "net_worth", 0.0)
                itr.long_term_debt = self._g(data, "long_term_debt", 0.0)
                itr.short_term_debt = self._g(data, "short_term_debt", 0.0)
                itr.revenue = self._g(data, "revenue", 0.0)
                itr.interest_expense = self._g(data, "interest_expense", 0.0)
                itr.depreciation = self._g(data, "depreciation", 0.0)
                itr.total_assets = self._g(data, "total_assets", 0.0)
                itr.total_liabilities = self._g(data, "total_liabilities", 0.0)

                if itr.net_income != 0 and (itr.interest_expense > 0 or itr.depreciation > 0):
                    itr.ebitda = (itr.net_income + itr.tax_paid
                                  + itr.interest_expense + itr.depreciation)

                if itr.net_worth > 0 or itr.gross_income > 0 or itr.revenue > 0:
                    if DEBUG_MODE:
                        print(f"ITR via Groq: net_worth={itr.net_worth}, "
                              f"revenue={itr.revenue}, LTD={itr.long_term_debt}, "
                              f"STD={itr.short_term_debt}, EBITDA={itr.ebitda}")
                    return itr

        print("ITR: Groq returned zeros — running regex fallback.")
        itr = self._extract_itr_regex(parsed)

        # If key balance-sheet fields are still missing, patch via openpyxl.
        # Condition: ANY of net_worth, long_term_debt, short_term_debt is zero
        print(
            f"ITR regex result: net_worth={itr.net_worth}, revenue={itr.revenue}")
        print(f"ITR source_file: {parsed.source_file!r}")
        if itr.net_worth == 0 or itr.long_term_debt == 0 or itr.short_term_debt == 0:
            print("ITR: regex incomplete — triggering openpyxl fallback...")
            itr = self._extract_itr_from_xlsx(parsed.source_file, itr)

        print(f"ITR final: net_worth={itr.net_worth}, revenue={itr.revenue}, "
              f"LTD={itr.long_term_debt}, STD={itr.short_term_debt}")
        return itr

    def _extract_itr_regex(self, parsed: ParsedDocument) -> ITRData:
        """
        Regex fallback for ITR extraction.
        Uses TWO strategies:
          1. Label: value  (colon-separated — works for PDFs and key:value layouts)
          2. Label  value  (whitespace-separated — works for XLSX/tabular layouts
             where Docling outputs "Revenue from Operations (Net of GST)  131200000")
        """
        text = parsed.raw_text
        itr = ITRData()

        # ── String fields ────────────────────────────────────────────────────
        for field, pattern in {
            'pan':             r'\b([A-Z]{5}[0-9]{4}[A-Z]{1})\b',
            'assessment_year': r'(?:Assessment Year|AY)[:\s]+(\d{4}-\d{2,4})',
        }.items():
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                setattr(itr, field, m.group(1).strip())

        # ── Numeric fields — three-tier: colon → pipe-table → whitespace ────────
        # Each tuple: (field, colon_pat, pipe_pat, whitespace_pat)
        # pipe_pat handles Docling's XLSX output: "| Label | | 123456789 | ... |"
        # Any pattern can be None to skip that tier.
        numeric_fields = [
            ('gross_income',
             r'(?:Gross Total Income|Total Income)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'(?:Gross Total Income|TOTAL INCOME)[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'(?:Gross Total Income|TOTAL INCOME)\s{2,}([\d,]+)'),
            ('net_income',
             r'(?:PROFIT AFTER TAX|PAT|Net Income)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'PROFIT AFTER TAX[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'PROFIT AFTER TAX.*?\s{2,}([\d,]+)'),
            ('tax_paid',
             r'(?:Tax Paid|Total Tax Paid|Provision for Current Tax)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'Provision for Current Tax[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'Provision for Current Tax.*?\s{2,}([\d,]+)'),
            ('tds',
             r'TDS[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             None, None),
            ('net_worth',
             r'(?:Net Worth|Networth|NET WORTH\s*/\s*EQUITY|Shareholders.{0,10}Equity)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'NET WORTH[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'NET WORTH.*?\s{2,}([\d,]+)'),
            ('long_term_debt',
             r'(?:Long.?term Borrowings?|Term Loan|Long.?term Debt)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'Long.?term Borrowings?[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'Long.?term Borrowings?.*?\s{2,}([\d,]+)'),
            ('short_term_debt',
             r'(?:Short.?term Borrowings?|Working Capital Loan|Cash Credit)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'Short.?term Borrowings?[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'Short.?term Borrowings?.*?\s{2,}([\d,]+)'),
            ('revenue',
             r'(?:Revenue from Operations|Net Revenue)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'Revenue from Operations[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'Revenue from Operations.*?\s{2,}([\d,]+)'),
            ('interest_expense',
             r'(?:Finance Costs?|Interest Expense|Interest Paid)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'Finance Costs?[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'Finance Costs?.*?\s{2,}([\d,]+)'),
            ('depreciation',
             r'(?:Depreciation|Amortis)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'Depreciation[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'Depreciation.*?\s{2,}([\d,]+)'),
            ('total_assets',
             r'(?:TOTAL ASSETS|Total Assets)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'TOTAL ASSETS[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'TOTAL ASSETS.*?\s{2,}([\d,]+)'),
            ('total_liabilities',
             r'(?:TOTAL (?:EQUITY\s*&\s*)?LIABILITIES|Total Liabilities)[:\s]*[₹Rs.]?\s*([\d,]+\.?\d*)',
             r'TOTAL EQUITY[^|\n]*LIABILITIES[^|\n]*\|[^|\n]*\|\s*([\d]{6,})',
             r'TOTAL EQUITY.*?LIABILITIES.*?\s{2,}([\d,]+)'),
        ]

        for field, colon_pat, pipe_pat, ws_pat in numeric_fields:
            if re.search(colon_pat, text, re.IGNORECASE):
                m = re.search(colon_pat, text, re.IGNORECASE)
                setattr(itr, field, self._parse_amount(m.group(1)))
            elif pipe_pat and re.search(pipe_pat, text, re.IGNORECASE):
                m = re.search(pipe_pat, text, re.IGNORECASE)
                setattr(itr, field, self._parse_amount(m.group(1)))
            elif ws_pat and re.search(ws_pat, text, re.IGNORECASE):
                m = re.search(ws_pat, text, re.IGNORECASE)
                setattr(itr, field, self._parse_amount(m.group(1)))

        if itr.net_income > 0 and itr.interest_expense > 0 and itr.depreciation > 0:
            itr.ebitda = (itr.net_income + itr.tax_paid
                          + itr.interest_expense + itr.depreciation)

        if DEBUG_MODE:
            print(f"ITR regex: net_worth={itr.net_worth}, revenue={itr.revenue}, "
                  f"LTD={itr.long_term_debt}, STD={itr.short_term_debt}")
        return itr

    def _extract_itr_from_xlsx(self, file_path: str, itr: "ITRData") -> "ITRData":
        """
        Direct openpyxl read for XLSX ITR/financial statements.
        Last-resort fallback when Docling text + regex both fail.
        Scans every cell for known label keywords and takes the first
        numeric value in the same row as the matched label.
        """
        try:
            import openpyxl
            from pathlib import Path as _Path
            import tempfile as _tempfile

            if not file_path:
                print("ITR xlsx fallback: no file_path provided")
                return itr

            p = _Path(file_path)

            # If path is relative (just a filename), check the system temp dir
            if not p.is_absolute():
                temp_candidate = _Path(_tempfile.gettempdir()) / p.name
                if temp_candidate.exists():
                    p = temp_candidate
                    file_path = str(p)

            if not p.exists():
                print(f"ITR xlsx fallback: file not found at {file_path!r}")
                return itr

            if not file_path.lower().endswith(('.xlsx', '.xls')):
                print(f"ITR xlsx fallback: not an xlsx file: {file_path!r}")
                return itr

            print(f"ITR xlsx fallback: reading {p}")

            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Map label keywords → ITRData field names
            # Each entry: (field, [keyword variants])
            label_map = [
                ('net_worth',        ['net worth', 'networth', 'equity']),
                ('revenue',          [
                 'revenue from operations', 'net revenue', 'net sales']),
                ('net_income',       [
                 'profit after tax', 'pat', 'net profit']),
                ('gross_income',     ['gross total income',
                 'total income', 'profit before tax']),
                ('tax_paid',         [
                 'provision for current tax', 'tax paid', 'income tax']),
                ('long_term_debt',   [
                 'long-term borrowings', 'long term borrowings', 'term loan']),
                ('short_term_debt',  ['short-term borrowings',
                 'short term borrowings', 'cash credit']),
                ('interest_expense', ['finance costs',
                 'interest expense', 'interest paid']),
                ('depreciation',     ['depreciation',
                 'amortisation', 'amortization']),
                ('total_assets',     ['total assets']),
                ('total_liabilities', [
                 'total equity & liabilities', 'total liabilities']),
            ]

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    # Build a text representation of the row
                    row_text = ' '.join(str(c).lower()
                                        for c in row if c is not None)
                    if not row_text.strip():
                        continue

                    for field, keywords in label_map:
                        if getattr(itr, field, 0) != 0:
                            continue  # Already filled
                        if any(kw in row_text for kw in keywords):
                            # Find the first numeric cell in this row
                            # Allow negatives for net_worth and net_income
                            # (negative net worth = insolvent; negative net income = loss)
                            allow_negative = field in (
                                'net_worth', 'net_income', 'gross_income')
                            for cell in row:
                                try:
                                    val = float(cell)
                                    if allow_negative and val != 0:
                                        setattr(itr, field, val)
                                        break
                                    elif not allow_negative and val > 0:
                                        setattr(itr, field, val)
                                        break
                                except (TypeError, ValueError):
                                    continue

            # Derive EBITDA if we now have the pieces
            if itr.net_income > 0 and (itr.interest_expense > 0 or itr.depreciation > 0):
                itr.ebitda = (itr.net_income + itr.tax_paid
                              + itr.interest_expense + itr.depreciation)

            print(f"ITR xlsx direct: net_worth={itr.net_worth}, revenue={itr.revenue}, "
                  f"LTD={itr.long_term_debt}, STD={itr.short_term_debt}")

        except Exception as e:
            print(f"ITR xlsx fallback error: {e}")
        return itr

    def extract(self, parsed: ParsedDocument):
        """Auto-detect document type and route to correct extractor."""
        doc_type = parsed.document_type
        if doc_type == DocumentType.GST_RETURN:
            return self.extract_gst(parsed)
        elif doc_type == DocumentType.BANK_STATEMENT:
            return self.extract_bank(parsed)
        elif doc_type == DocumentType.ITR:
            return self.extract_itr(parsed)
        else:
            gst = self.extract_gst(parsed)
            bank = self.extract_bank(parsed)
            itr = self.extract_itr(parsed)
            gst_score = sum(
                1 for v in [gst.gstin, gst.turnover, gst.itc_claimed] if v)
            bank_score = sum(
                1 for v in [bank.total_credits, bank.total_debits] if v)
            itr_score = sum(1 for v in [itr.gross_income, itr.net_worth] if v)
            if gst_score >= bank_score and gst_score >= itr_score:
                return gst
            elif bank_score >= itr_score:
                return bank
            return itr

    # ─── HELPERS ─────────────────────────────────────────────────────────────

    def _parse_amount(self, value: str) -> float:
        """Convert Indian number format to float."""
        if not value:
            return 0.0
        cleaned = re.sub(r'[₹Rs.\s]', '', str(value))
        cleaned_lower = cleaned.lower()
        if 'cr' in cleaned_lower:
            num = re.sub(r'[^0-9.]', '', cleaned)
            return float(num) * 10_000_000 if num else 0.0
        if 'l' in cleaned_lower or 'lakh' in cleaned_lower:
            num = re.sub(r'[^0-9.]', '', cleaned)
            return float(num) * 100_000 if num else 0.0
        cleaned = cleaned.replace(',', '')
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return 0.0


if __name__ == "__main__":
    from src.parser import DocumentParser

    parser = DocumentParser()
    extractor = FinancialExtractor()

    print(f"Groq available: {extractor._groq_available}")

    test_file = input("Enter path to a GST/Bank/ITR file to test: ").strip()
    if test_file and Path(test_file).exists():
        print("\nParsing...")
        parsed = parser.parse(test_file)
        print("Extracting...")
        result = extractor.extract(parsed)
        print(f"\n--- EXTRACTED ({type(result).__name__}) ---")
        for k, v in result.model_dump().items():
            if v and v != 0 and v != [] and v != "":
                print(f"  {k}: {v}")
    else:
        print("No file provided.")
